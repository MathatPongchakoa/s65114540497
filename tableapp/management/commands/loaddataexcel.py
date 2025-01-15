import os
from django.core.files import File
from openpyxl import load_workbook
from django.conf import settings
from tableapp.models import *
from datetime import datetime
from django.core.management.base import BaseCommand
from io import BytesIO
from django.core.files.base import ContentFile


class Command(BaseCommand):
    help = 'Load data from Excel into CustomUser, Table, Booking, Category, and Menu models'

    def handle(self, *args, **options):
        file_path = os.path.join(settings.BASE_DIR, 'tableapp', 'fixtures', 'data_ex.xlsx')

        # อ่านข้อมูลจาก Excel
        wb = load_workbook(file_path)

        # แผ่นงานสำหรับ CustomUser
        if 'CustomUser' in wb.sheetnames:
            user_sheet = wb['CustomUser']
            for row in user_sheet.iter_rows(min_row=2, values_only=True):
                user_id, username, email, first_name, last_name, password = row
                if password and not isinstance(password, str):
                    password = str(password)
                if password:
                    user_instance, created = CustomUser.objects.get_or_create(
                        id=user_id,
                        defaults={
                            'username': username,
                            'email': email,
                            'first_name': first_name,
                            'last_name': last_name,
                        }
                    )
                    if created:
                        user_instance.set_password(password)
                        user_instance.save()
                        self.stdout.write(f"Created user: {username}")
                    else:
                        self.stdout.write(f"User already exists: {username}")
                else:
                    self.stdout.write(f"Skipping user {username} due to missing password")

        if 'Zone' in wb.sheetnames:
            zone_sheet = wb['Zone']
            for row in zone_sheet.iter_rows(min_row=2, values_only=True):
                name, description, image_path = row[1:4]
                self.stdout.write(f"Processing zone: {name}, description: {description}, image_path: {image_path}")

                # ตรวจสอบ path ของไฟล์ภาพ
                image_file = None
                if image_path:
                    full_image_path = os.path.join(settings.MEDIA_ROOT, 'zone_images', image_path)
                    self.stdout.write(f"Checking image path: {full_image_path}")  # Debug path
                    if os.path.exists(full_image_path):  # ตรวจสอบว่าไฟล์ภาพมีอยู่จริง
                        self.stdout.write(f"Image found: {full_image_path}")
                        try:
                            # อ่านไฟล์ภาพและสร้าง ContentFile
                            with open(full_image_path, 'rb') as img_file:
                                image_content = ContentFile(img_file.read(), name=image_path)
                                self.stdout.write(f"File object created for image: {image_path}")
                        except Exception as e:
                            self.stdout.write(f"Error reading image file '{image_path}': {str(e)}")
                            continue
                    else:
                        self.stdout.write(f"Image not found: {full_image_path}")

                # สร้างหรืออัปเดต Zone
                try:
                    zone_instance, created = Zone.objects.get_or_create(
                        name=name,
                        defaults={'description': description}
                    )

                    if created:
                        self.stdout.write(f"Created zone: {name}")
                    else:
                        self.stdout.write(f"Zone already exists: {name}")

                    # ตรวจสอบและบันทึกภาพ
                    if image_content:
                        # ตรวจสอบว่าภาพที่อยู่ในฐานข้อมูลตรงกับภาพใหม่หรือไม่
                        if not zone_instance.image or not zone_instance.image.name.endswith(os.path.basename(image_path)):
                            zone_instance.image.save(os.path.basename(image_path), image_content, save=True)
                            self.stdout.write(f"Image saved for zone: {name}")
                        else:
                            self.stdout.write(f"Image for zone '{name}' already up-to-date.")

                except Exception as e:
                    self.stdout.write(f"Error processing zone '{name}': {str(e)}")




        # แผ่นงานสำหรับ Table
        if 'Table' in wb.sheetnames:
            table_sheet = wb['Table']
            for row in table_sheet.iter_rows(min_row=2, values_only=True):
                table_id, table_name, table_status, zone_id = row  # เพิ่มการอ่าน zone_id

                # ตรวจสอบว่าค่า table_status ถูกต้อง
                valid_statuses = ['available', 'occupied', 'booked']
                if table_status not in valid_statuses:
                    self.stdout.write(f"Invalid table status '{table_status}' for table {table_name}. Skipping...")
                    continue

                # ค้นหา Zone ด้วย zone_id
                zone = None
                if zone_id:
                    zone = Zone.objects.filter(id=zone_id).first()
                    if not zone:
                        self.stdout.write(f"Zone ID {zone_id} not found for table {table_name}. Skipping...")
                        continue

                # สร้างหรืออัปเดต Table
                table_instance, created = Table.objects.get_or_create(
                    id=table_id,
                    defaults={
                        'table_name': table_name,
                        'table_status': table_status,
                        'zone': zone,  # เพิ่ม zone ในการสร้าง
                    }
                )
                if created:
                    self.stdout.write(f"Created table: {table_name} in zone {zone.name if zone else 'N/A'}")
                else:
                    # อัปเดต zone ถ้ามีการเปลี่ยนแปลง
                    table_instance.zone = zone
                    table_instance.table_status = table_status
                    table_instance.save()
                    self.stdout.write(f"Updated table: {table_name} in zone {zone.name if zone else 'N/A'}")

        # แผ่นงานสำหรับ Booking
        if 'Booking' in wb.sheetnames:
            booking_sheet = wb['Booking']
            for row in booking_sheet.iter_rows(min_row=2, values_only=True):
                booking_id, table_id, booking_date, booking_time, booking_end_time, user_id, status = row

                # ตรวจสอบวันที่
                try:
                    booking_date = datetime.strptime(str(booking_date), '%Y-%m-%d').date()
                except ValueError as e:
                    self.stdout.write(f"Invalid booking date format for booking ID {booking_id}. Skipping... Error: {e}")
                    continue

                # ตรวจสอบเวลา
                try:
                    booking_time = datetime.strptime(str(booking_time), '%H:%M:%S').time()
                    booking_end_time = datetime.strptime(str(booking_end_time), '%H:%M:%S').time()
                except ValueError as e:
                    self.stdout.write(f"Invalid booking time format for booking ID {booking_id}. Skipping... Error: {e}")
                    continue

                # ค้นหา Table
                table = Table.objects.filter(id=table_id).first()
                if not table:
                    self.stdout.write(f"Table ID {table_id} not found. Skipping booking.")
                    continue

                # ค้นหา User
                user = None
                if user_id:
                    user = CustomUser.objects.filter(id=user_id).first()
                    if not user:
                        self.stdout.write(f"User ID {user_id} not found. Skipping booking ID {booking_id}.")

                # ตรวจสอบสถานะ booking
                valid_statuses = ['pending', 'occupied', 'cancelled']
                if status not in valid_statuses:
                    self.stdout.write(f"Invalid booking status '{status}' for booking ID {booking_id}. Skipping...")
                    continue

                # สร้างข้อมูลการจอง
                booking_instance, created = Booking.objects.get_or_create(
                    id=booking_id,
                    defaults={
                        'table': table,
                        'booking_date': booking_date,
                        'booking_time': booking_time,
                        'booking_end_time': booking_end_time,
                        'user': user,
                        'status': status,
                    }
                )
                if created:
                    self.stdout.write(f"Created booking ID {booking_id} for table {table.table_name}")
                else:
                    self.stdout.write(f"Booking ID {booking_id} already exists.")

        # แผ่นงานสำหรับ Category
        if 'Category' in wb.sheetnames:
            category_sheet = wb['Category']
            for row in category_sheet.iter_rows(min_row=2, values_only=True):
                name = row[1]
                category_instance, created = Category.objects.get_or_create(name=name)
                if created:
                    self.stdout.write(f"Created category: {name}")
                else:
                    self.stdout.write(f"Category already exists: {name}")

                # แผ่นงานสำหรับ Menu
        # แผ่นงานสำหรับ Menu
        if 'Menu' in wb.sheetnames:
            menu_sheet = wb['Menu']
            for row in menu_sheet.iter_rows(min_row=2, values_only=True):
                food_name, price, image_path, category_name = row[1:5]

                # ตรวจสอบหรือสร้าง Category
                category, _ = Category.objects.get_or_create(name=category_name)

                # ค้นหาหรือสร้าง Menu โดยไม่มี image ก่อน
                menu_instance, created = Menu.objects.get_or_create(
                        food_name=food_name,
                        defaults={
                            'price': price,
                            'category': category,
                            'image': image_path,  # เพิ่มตรงนี้
                        }
                    )

                # ตรวจสอบว่าเมนูยังไม่มี image และ image_path ถูกต้อง
                if image_path:
                    full_image_path = os.path.join(settings.MEDIA_ROOT, image_path)
                    print(f"Checking image path: {full_image_path}")  # Debug path
                    if os.path.exists(full_image_path):
                        try:
                            # อัปเดตเฉพาะกรณีที่ไม่มีภาพเดิม
                            if not menu_instance.image or not menu_instance.image.name.endswith(image_path):
                                with open(full_image_path, 'rb') as img_file:
                                    django_file = File(img_file)
                                    menu_instance.image.save(os.path.basename(image_path), django_file, save=True)
                                    print(f"Image saved for menu: {food_name}")
                        except Exception as e:
                            print(f"Error saving image for menu '{food_name}': {str(e)}")
                    else:
                        print(f"Image not found: {full_image_path}")



                if created:
                    print(f"Created menu item: {food_name}")
                else:
                    print(f"Menu item already exists: {food_name}")






        self.stdout.write(self.style.SUCCESS("Data loaded successfully!"))
